from openpyxl import load_workbook
from openpyxl.cell import MergedCell
from openpyxl.drawing.image import Image
from models import Quote
import os
import shutil
from copy import copy
from datetime import datetime

def is_writable(ws, row, col):
    return not isinstance(ws.cell(row=row, column=col), MergedCell)

def copy_style(src_cell, dest_cell):
    dest_cell.font = copy(src_cell.font)
    dest_cell.border = copy(src_cell.border)
    dest_cell.fill = copy(src_cell.fill)
    dest_cell.number_format = copy(src_cell.number_format)
    dest_cell.protection = copy(src_cell.protection)
    dest_cell.alignment = copy(src_cell.alignment)

def save_quote_to_excel(quote: Quote):
    # Use the original, unmodified template
    template_path = os.path.join("templates", "ALUMINIUM PARTITION QUOTE.xlsx")
    output_dir = "output"
    os.makedirs(output_dir, exist_ok=True)

    # Create versioned output filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = quote.customer.name.replace(" ", "_").replace("/", "_")
    output_path = os.path.join(output_dir, f"Quote_{safe_name}_{timestamp}.xlsx")

    # Copy the template without changing it
    shutil.copy(template_path, output_path)
    wb = load_workbook(output_path)
    ws = wb.active

    # Insert company logo if exists (top-right)
    logo_path = os.path.join("templates", "company.png")
    if os.path.exists(logo_path):
        img = Image(logo_path)
        img.width = 160
        img.height = 80
        ws.add_image(img, "L1")

    # Customer and quotation fields
    fields = {
        "D17": quote.customer.name,
        "D18": quote.customer.address,
        "D19": quote.customer.town,
        "D20": quote.customer.contact,
        "F20": quote.customer.email,
        "K16": quote.quote.quotation_no,
        "K17": quote.quote.date,
        "D24": quote.quote.service,
    }

    for cell, value in fields.items():
        row, col = ws[cell].row, ws[cell].column
        if is_writable(ws, row, col):
            ws[cell].value = value

    # Add item entries
    start_row = 27
    template_row = 27
    for i, item in enumerate(quote.items):
        row = start_row + i
        item_cells = {
            f"A{row}": item.qty,
            f"B{row}": item.description,
            f"E{row}": item.units,
            f"F{row}": item.unit_price,
            f"K{row}": item.total,
        }
        for cell, value in item_cells.items():
            r, c = ws[cell].row, ws[cell].column
            if is_writable(ws, r, c):
                ws[cell].value = value
                copy_style(ws[f"{cell[0]}{template_row}"], ws[cell])

    # Labour row
    labour_row = start_row + len(quote.items)
    labour_cells = {
        f"B{labour_row}": quote.labour.description,
        f"E{labour_row}": quote.labour.units,
        f"K{labour_row}": quote.labour.total,
    }
    for cell, value in labour_cells.items():
        r, c = ws[cell].row, ws[cell].column
        if is_writable(ws, r, c):
            ws[cell].value = value
            copy_style(ws[f"{cell[0]}{template_row}"], ws[cell])

    # Total row (a few rows below labour)
    total_row = labour_row + 2
    total_cell = f"K{total_row}"
    if is_writable(ws, total_row, 11):  # Column 11 is 'K'
        ws[total_cell].value = quote.grand_total
        copy_style(ws[f"K{template_row}"], ws[total_cell])

    wb.save(output_path)
    print(f"✅ Quote saved to: {output_path}")

